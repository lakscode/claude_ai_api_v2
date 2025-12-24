"""
Output Generator Module
Generates Excel and PDF output files for lease clause classification results.
"""

import os
from pathlib import Path
from datetime import datetime


def generate_excel_output(results, output_folder, filename_prefix="lease_classification"):
    """
    Generate Excel file from classification results.

    Args:
        results: List of classification result dictionaries.
        output_folder: Path to output folder.
        filename_prefix: Prefix for output filename.

    Returns:
        Path to generated Excel file or None if failed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        output_path = Path(output_folder)
        output_path.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = output_path / f"{filename_prefix}_{timestamp}.xlsx"

        wb = Workbook()

        # Create Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Summary headers
        summary_headers = ["PDF File", "Total Clauses", "Total Fields", "API Calls", "Processing Status"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Summary data
        for row, result in enumerate(results, 2):
            ws_summary.cell(row=row, column=1, value=result.get("pdf_file", "")).border = border
            ws_summary.cell(row=row, column=2, value=result.get("total_clauses", 0)).border = border
            ws_summary.cell(row=row, column=3, value=result.get("total_fields", 0)).border = border
            ws_summary.cell(row=row, column=4, value=result.get("openai_api_calls", 0)).border = border
            ws_summary.cell(row=row, column=5, value="Success" if result else "Failed").border = border

        # Adjust column widths for summary
        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 15
        ws_summary.column_dimensions['D'].width = 12
        ws_summary.column_dimensions['E'].width = 18

        # Create Clauses sheet
        ws_clauses = wb.create_sheet("Clauses")
        clause_headers = ["PDF File", "Clause Index", "Clause Type", "Type ID", "Confidence", "Clause Text"]
        for col, header in enumerate(clause_headers, 1):
            cell = ws_clauses.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        clause_row = 2
        for result in results:
            pdf_file = result.get("pdf_file", "")
            for clause in result.get("clauses", []):
                ws_clauses.cell(row=clause_row, column=1, value=pdf_file).border = border
                ws_clauses.cell(row=clause_row, column=2, value=clause.get("clause_index", "")).border = border
                ws_clauses.cell(row=clause_row, column=3, value=clause.get("type", "")).border = border
                ws_clauses.cell(row=clause_row, column=4, value=clause.get("type_id", "")).border = border
                ws_clauses.cell(row=clause_row, column=5, value=clause.get("confidence", 0)).border = border

                # Truncate clause text if too long for Excel cell
                clause_text = clause.get("text", "")
                if len(clause_text) > 32000:
                    clause_text = clause_text[:32000] + "..."
                cell = ws_clauses.cell(row=clause_row, column=6, value=clause_text)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True)

                clause_row += 1

        # Adjust column widths for clauses
        ws_clauses.column_dimensions['A'].width = 30
        ws_clauses.column_dimensions['B'].width = 12
        ws_clauses.column_dimensions['C'].width = 25
        ws_clauses.column_dimensions['D'].width = 25
        ws_clauses.column_dimensions['E'].width = 12
        ws_clauses.column_dimensions['F'].width = 80

        # Create Fields sheet
        ws_fields = wb.create_sheet("Fields")
        field_headers = ["PDF File", "Field Name", "Field ID", "Values", "Clause Indices"]
        for col, header in enumerate(field_headers, 1):
            cell = ws_fields.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        field_row = 2
        for result in results:
            pdf_file = result.get("pdf_file", "")
            for field in result.get("fields", []):
                ws_fields.cell(row=field_row, column=1, value=pdf_file).border = border
                ws_fields.cell(row=field_row, column=2, value=field.get("field_name", "")).border = border
                ws_fields.cell(row=field_row, column=3, value=field.get("field_id", "")).border = border

                # Join values array into comma-separated string
                values = field.get("values", [])
                values_str = ", ".join(str(v) for v in values) if values else ""
                ws_fields.cell(row=field_row, column=4, value=values_str).border = border

                # Join clause indices
                indices = field.get("clause_indices", [])
                indices_str = ", ".join(str(i) for i in indices) if indices else ""
                ws_fields.cell(row=field_row, column=5, value=indices_str).border = border

                field_row += 1

        # Adjust column widths for fields
        ws_fields.column_dimensions['A'].width = 30
        ws_fields.column_dimensions['B'].width = 30
        ws_fields.column_dimensions['C'].width = 25
        ws_fields.column_dimensions['D'].width = 50
        ws_fields.column_dimensions['E'].width = 20

        wb.save(excel_file)
        print(f"Excel output saved: {excel_file}")
        return str(excel_file)

    except ImportError as e:
        print(f"Error: openpyxl not installed. Run: pip install openpyxl")
        return None
    except Exception as e:
        print(f"Error generating Excel output: {str(e)}")
        return None


def generate_pdf_output(results, output_folder, filename_prefix="lease_classification"):
    """
    Generate PDF file from classification results.

    Args:
        results: List of classification result dictionaries.
        output_folder: Path to output folder.
        filename_prefix: Prefix for output filename.

    Returns:
        Path to generated PDF file or None if failed.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_LEFT, TA_CENTER

        output_path = Path(output_folder)
        output_path.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_file = output_path / f"{filename_prefix}_{timestamp}.pdf"

        doc = SimpleDocTemplate(
            str(pdf_file),
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            spaceBefore=15
        )
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            leading=12
        )
        small_style = ParagraphStyle(
            'Small',
            parent=styles['Normal'],
            fontSize=8,
            leading=10
        )

        elements = []

        # Title
        elements.append(Paragraph("Lease Clause Classification Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
        elements.append(Spacer(1, 20))

        # Summary Section
        elements.append(Paragraph("Summary", heading_style))

        summary_data = [["PDF File", "Clauses", "Fields", "API Calls", "Status"]]
        for result in results:
            summary_data.append([
                result.get("pdf_file", "")[:40],
                str(result.get("total_clauses", 0)),
                str(result.get("total_fields", 0)),
                str(result.get("openai_api_calls", 0)),
                "Success" if result else "Failed"
            ])

        summary_table = Table(summary_data, colWidths=[180, 60, 60, 60, 60])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E8EEF7')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Style for field table cells
        field_cell_style = ParagraphStyle(
            'FieldCellText',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            wordWrap='CJK'
        )

        # Extracted Fields Section
        elements.append(Paragraph("Extracted Fields", heading_style))

        for result in results:
            pdf_name = result.get("pdf_file", "Unknown")
            fields = result.get("fields", [])

            if fields:
                elements.append(Paragraph(f"<b>{pdf_name}</b>", normal_style))
                elements.append(Spacer(1, 5))

                field_data = [[
                    Paragraph("<b>Field Name</b>", field_cell_style),
                    Paragraph("<b>Values</b>", field_cell_style)
                ]]
                for field in fields:
                    field_name = field.get("field_name", "")
                    values = field.get("values", [])
                    values_str = ", ".join(str(v) for v in values)
                    # Escape special characters for Paragraph
                    values_str = values_str.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    field_name = field_name.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    field_data.append([
                        Paragraph(field_name, field_cell_style),
                        Paragraph(values_str, field_cell_style)
                    ])

                field_table = Table(field_data, colWidths=[150, 350])
                field_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5B9BD5')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#DEEBF7')),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                elements.append(field_table)
                elements.append(Spacer(1, 15))

        # Style for wrapped text in table cells
        cell_style = ParagraphStyle(
            'CellText',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
            wordWrap='CJK'
        )
        cell_style_type = ParagraphStyle(
            'CellType',
            parent=styles['Normal'],
            fontSize=7,
            leading=9
        )

        # Clauses Section (new page)
        elements.append(PageBreak())
        elements.append(Paragraph("Classified Clauses", heading_style))

        for result in results:
            pdf_name = result.get("pdf_file", "Unknown")
            clauses = result.get("clauses", [])

            if clauses:
                elements.append(Paragraph(f"<b>{pdf_name}</b> ({len(clauses)} clauses)", normal_style))
                elements.append(Spacer(1, 5))

                # Header row with plain text
                clause_data = [[
                    Paragraph("<b>#</b>", cell_style),
                    Paragraph("<b>Type</b>", cell_style),
                    Paragraph("<b>Confidence</b>", cell_style),
                    Paragraph("<b>Text Preview</b>", cell_style)
                ]]
                for clause in clauses[:50]:  # Limit to first 50 clauses per PDF
                    clause_text = clause.get("text", "")
                    # Truncate text for PDF but allow more text with wrapping
                    if len(clause_text) > 300:
                        clause_text = clause_text[:300] + "..."
                    # Escape special characters for Paragraph
                    clause_text = clause_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    clause_type = clause.get("type", "")[:30]

                    clause_data.append([
                        Paragraph(str(clause.get("clause_index", "")), cell_style),
                        Paragraph(clause_type, cell_style_type),
                        Paragraph(f"{clause.get('confidence', 0):.2%}", cell_style),
                        Paragraph(clause_text, cell_style)
                    ])

                clause_table = Table(clause_data, colWidths=[25, 90, 50, 340])
                clause_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                    ('ALIGN', (2, 1), (2, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E2EFDA')),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                elements.append(clause_table)

                if len(clauses) > 50:
                    elements.append(Paragraph(f"<i>... and {len(clauses) - 50} more clauses</i>", small_style))

                elements.append(Spacer(1, 15))

        doc.build(elements)
        print(f"PDF output saved: {pdf_file}")
        return str(pdf_file)

    except ImportError as e:
        print(f"Error: reportlab not installed. Run: pip install reportlab")
        return None
    except Exception as e:
        print(f"Error generating PDF output: {str(e)}")
        return None


def generate_outputs(results, output_folder="output_files", filename_prefix="lease_classification"):
    """
    Generate both Excel and PDF outputs.

    Args:
        results: List of classification result dictionaries.
        output_folder: Path to output folder.
        filename_prefix: Prefix for output filenames.

    Returns:
        Dictionary with paths to generated files.
    """
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)

    generated_files = {
        "excel": None,
        "pdf": None
    }

    # Generate Excel
    excel_path = generate_excel_output(results, output_folder, filename_prefix)
    if excel_path:
        generated_files["excel"] = excel_path

    # Generate PDF
    pdf_path = generate_pdf_output(results, output_folder, filename_prefix)
    if pdf_path:
        generated_files["pdf"] = pdf_path

    return generated_files
