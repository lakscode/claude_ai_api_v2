"""
Data retrieval routes for the Lease Clause Classifier API.
Handles fetching, searching, deleting, and exporting classification data from MongoDB.
"""

import os
import io
import json
import tempfile
from datetime import datetime

from flask import Blueprint, request, jsonify, current_app, send_file, Response

from utils import log_success, log_error
from db import (
    get_mongo_client,
    get_mongo_config,
    find_document_by_id,
    delete_document_by_id,
    serialize_document
)

data_bp = Blueprint('data', __name__)


@data_bp.route('/data', methods=['GET'])
def get_all_data():
    """
    Get all stored classification data from MongoDB.

    Query Parameters:
        - limit: Maximum number of records to return (default: 100, max: 1000)
        - skip: Number of records to skip for pagination (default: 0)
        - sort: Sort order - 'asc' or 'desc' by created_at (default: desc)

    Response:
        JSON with list of all classification results and pagination info.
    """
    try:
        config = current_app.config.get('APP_CONFIG', {})
        mongo_uri, mongo_db, mongo_collection = get_mongo_config(config)

        if not mongo_uri or not mongo_db:
            log_error("MongoDB not configured", endpoint="/data")
            return jsonify({"error": "MongoDB not configured"}), 500

        # Get pagination parameters
        limit = min(int(request.args.get('limit', 100)), 1000)
        skip = int(request.args.get('skip', 0))
        sort_order = request.args.get('sort', 'desc')
        sort_direction = -1 if sort_order == 'desc' else 1

        from pymongo import MongoClient

        client = MongoClient(mongo_uri)
        db = client[mongo_db]
        collection = db[mongo_collection]

        # Get total count
        total_count = collection.count_documents({})

        # Fetch data with pagination
        cursor = collection.find({}).sort("created_at", sort_direction).skip(skip).limit(limit)

        results = []
        for doc in cursor:
            results.append(serialize_document(doc))

        client.close()

        log_success("Data retrieved from MongoDB", endpoint="/data", count=len(results), total=total_count)

        return jsonify({
            "total": total_count,
            "limit": limit,
            "skip": skip,
            "count": len(results),
            "data": results
        }), 200

    except ImportError as e:
        log_error("pymongo library not installed", endpoint="/data", error=str(e))
        return jsonify({"error": "pymongo library not installed"}), 500
    except Exception as e:
        log_error("Failed to retrieve data", endpoint="/data", error=str(e))
        return jsonify({"error": str(e)}), 500


@data_bp.route('/data/<doc_id>', methods=['GET'])
def get_data_by_id(doc_id):
    """
    Get a specific classification result by document ID.

    Path Parameters:
        - doc_id: MongoDB document ID

    Response:
        JSON with the classification result.
    """
    try:
        config = current_app.config.get('APP_CONFIG', {})
        mongo_uri, mongo_db, mongo_collection = get_mongo_config(config)

        if not mongo_uri or not mongo_db:
            log_error("MongoDB not configured", endpoint=f"/data/{doc_id}")
            return jsonify({"error": "MongoDB not configured"}), 500

        from pymongo import MongoClient

        client = MongoClient(mongo_uri)
        db = client[mongo_db]
        collection = db[mongo_collection]

        doc = find_document_by_id(collection, doc_id)
        client.close()

        if not doc:
            log_error("Document not found", endpoint=f"/data/{doc_id}", doc_id=doc_id)
            return jsonify({"error": "Document not found"}), 404

        doc = serialize_document(doc)

        log_success("Document retrieved", endpoint=f"/data/{doc_id}", doc_id=doc_id)

        return jsonify(doc), 200

    except ImportError as e:
        log_error("pymongo library not installed", endpoint=f"/data/{doc_id}", error=str(e))
        return jsonify({"error": "pymongo library not installed"}), 500
    except Exception as e:
        log_error("Failed to retrieve document", endpoint=f"/data/{doc_id}", error=str(e))
        return jsonify({"error": str(e)}), 500


@data_bp.route('/data/search', methods=['GET'])
def search_data():
    """
    Search classification data by PDF filename or field values.

    Query Parameters:
        - filename: Search by PDF filename (partial match)
        - field_name: Search by field name
        - field_value: Search by field value (use with field_name)
        - limit: Maximum number of records (default: 100)

    Response:
        JSON with matching classification results.
    """
    try:
        config = current_app.config.get('APP_CONFIG', {})
        mongo_uri, mongo_db, mongo_collection = get_mongo_config(config)

        if not mongo_uri or not mongo_db:
            log_error("MongoDB not configured", endpoint="/data/search")
            return jsonify({"error": "MongoDB not configured"}), 500

        # Get search parameters
        filename = request.args.get('filename', '')
        field_name = request.args.get('field_name', '')
        field_value = request.args.get('field_value', '')
        limit = min(int(request.args.get('limit', 100)), 1000)

        from pymongo import MongoClient

        client = MongoClient(mongo_uri)
        db = client[mongo_db]
        collection = db[mongo_collection]

        # Build query
        query = {}
        if filename:
            query['pdf_file'] = {'$regex': filename, '$options': 'i'}
        if field_name:
            query['fields.field_name'] = {'$regex': field_name, '$options': 'i'}
        if field_value:
            query['fields.values'] = {'$regex': field_value, '$options': 'i'}

        if not query:
            client.close()
            return jsonify({"error": "At least one search parameter required (filename, field_name, or field_value)"}), 400

        # Fetch matching documents
        cursor = collection.find(query).sort("created_at", -1).limit(limit)

        results = []
        for doc in cursor:
            results.append(serialize_document(doc))

        client.close()

        log_success("Search completed", endpoint="/data/search", query=str(query), count=len(results))

        return jsonify({
            "count": len(results),
            "query": {
                "filename": filename or None,
                "field_name": field_name or None,
                "field_value": field_value or None
            },
            "data": results
        }), 200

    except ImportError as e:
        log_error("pymongo library not installed", endpoint="/data/search", error=str(e))
        return jsonify({"error": "pymongo library not installed"}), 500
    except Exception as e:
        log_error("Search failed", endpoint="/data/search", error=str(e))
        return jsonify({"error": str(e)}), 500


@data_bp.route('/data/<doc_id>', methods=['DELETE'])
def delete_data(doc_id):
    """
    Delete a specific classification result by document ID.

    Path Parameters:
        - doc_id: MongoDB document ID

    Response:
        JSON with deletion status.
    """
    try:
        config = current_app.config.get('APP_CONFIG', {})
        mongo_uri, mongo_db, mongo_collection = get_mongo_config(config)

        if not mongo_uri or not mongo_db:
            log_error("MongoDB not configured", endpoint=f"/data/{doc_id}")
            return jsonify({"error": "MongoDB not configured"}), 500

        from pymongo import MongoClient

        client = MongoClient(mongo_uri)
        db = client[mongo_db]
        collection = db[mongo_collection]

        deleted_count = delete_document_by_id(collection, doc_id)
        client.close()

        if deleted_count == 0:
            log_error("Document not found for deletion", endpoint=f"/data/{doc_id}", doc_id=doc_id)
            return jsonify({"error": "Document not found"}), 404

        log_success("Document deleted", endpoint=f"/data/{doc_id}", doc_id=doc_id)

        return jsonify({
            "message": "Document deleted successfully",
            "doc_id": doc_id
        }), 200

    except ImportError as e:
        log_error("pymongo library not installed", endpoint=f"/data/{doc_id}", error=str(e))
        return jsonify({"error": "pymongo library not installed"}), 500
    except Exception as e:
        log_error("Failed to delete document", endpoint=f"/data/{doc_id}", error=str(e))
        return jsonify({"error": str(e)}), 500


@data_bp.route('/data/stats', methods=['GET'])
def get_data_stats():
    """
    Get statistics about stored classification data.

    Response:
        JSON with database statistics.
    """
    try:
        config = current_app.config.get('APP_CONFIG', {})
        mongo_uri, mongo_db, mongo_collection = get_mongo_config(config)

        if not mongo_uri or not mongo_db:
            log_error("MongoDB not configured", endpoint="/data/stats")
            return jsonify({"error": "MongoDB not configured"}), 500

        from pymongo import MongoClient

        client = MongoClient(mongo_uri)
        db = client[mongo_db]
        collection = db[mongo_collection]

        # Get statistics
        total_documents = collection.count_documents({})

        # Aggregate statistics
        pipeline = [
            {
                "$group": {
                    "_id": None,
                    "total_clauses": {"$sum": "$total_clauses"},
                    "total_fields": {"$sum": "$total_fields"},
                    "total_api_calls": {"$sum": "$openai_api_calls"},
                    "avg_clauses": {"$avg": "$total_clauses"},
                    "avg_fields": {"$avg": "$total_fields"}
                }
            }
        ]

        stats_result = list(collection.aggregate(pipeline))

        # Get unique PDF files
        unique_pdfs = len(collection.distinct("pdf_file"))

        # Get date range
        oldest = collection.find_one({}, sort=[("created_at", 1)])
        newest = collection.find_one({}, sort=[("created_at", -1)])

        client.close()

        stats = {
            "total_documents": total_documents,
            "unique_pdfs": unique_pdfs,
            "database": mongo_db,
            "collection": mongo_collection
        }

        if stats_result:
            agg = stats_result[0]
            stats["total_clauses_processed"] = agg.get("total_clauses", 0)
            stats["total_fields_extracted"] = agg.get("total_fields", 0)
            stats["total_api_calls"] = agg.get("total_api_calls", 0)
            stats["avg_clauses_per_doc"] = round(agg.get("avg_clauses", 0), 2)
            stats["avg_fields_per_doc"] = round(agg.get("avg_fields", 0), 2)

        if oldest and 'created_at' in oldest:
            stats["oldest_record"] = oldest['created_at'].isoformat() if hasattr(oldest['created_at'], 'isoformat') else str(oldest['created_at'])
        if newest and 'created_at' in newest:
            stats["newest_record"] = newest['created_at'].isoformat() if hasattr(newest['created_at'], 'isoformat') else str(newest['created_at'])

        log_success("Stats retrieved", endpoint="/data/stats", total=total_documents)

        return jsonify(stats), 200

    except ImportError as e:
        log_error("pymongo library not installed", endpoint="/data/stats", error=str(e))
        return jsonify({"error": "pymongo library not installed"}), 500
    except Exception as e:
        log_error("Failed to retrieve stats", endpoint="/data/stats", error=str(e))
        return jsonify({"error": str(e)}), 500


@data_bp.route('/data/export/json', methods=['GET'])
def export_json():
    """
    Export classification data as JSON file.

    Query Parameters:
        - limit: Maximum number of records to export (default: 1000)
        - doc_id: Export specific document by ID (optional)

    Response:
        JSON file download.
    """
    try:
        config = current_app.config.get('APP_CONFIG', {})
        mongo_uri, mongo_db, mongo_collection = get_mongo_config(config)

        if not mongo_uri or not mongo_db:
            log_error("MongoDB not configured", endpoint="/data/export/json")
            return jsonify({"error": "MongoDB not configured"}), 500

        from pymongo import MongoClient

        client = MongoClient(mongo_uri)
        db = client[mongo_db]
        collection = db[mongo_collection]

        # Check for specific document
        doc_id = request.args.get('doc_id')
        if doc_id:
            doc = find_document_by_id(collection, doc_id)
            client.close()
            if not doc:
                return jsonify({"error": "Document not found"}), 404
            results = [serialize_document(doc)]
        else:
            limit = min(int(request.args.get('limit', 1000)), 10000)
            cursor = collection.find({}).sort("created_at", -1).limit(limit)
            results = [serialize_document(doc) for doc in cursor]
            client.close()

        # Create JSON response
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"lease_classification_{timestamp}.json"

        json_data = json.dumps(results, indent=2, ensure_ascii=False, default=str)

        log_success("JSON export completed", endpoint="/data/export/json", count=len(results))

        return Response(
            json_data,
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except ImportError as e:
        log_error("pymongo library not installed", endpoint="/data/export/json", error=str(e))
        return jsonify({"error": "pymongo library not installed"}), 500
    except Exception as e:
        log_error("JSON export failed", endpoint="/data/export/json", error=str(e))
        return jsonify({"error": str(e)}), 500


@data_bp.route('/data/export/excel', methods=['GET'])
def export_excel():
    """
    Export classification data as Excel file.

    Query Parameters:
        - limit: Maximum number of records to export (default: 1000)
        - doc_id: Export specific document by ID (optional)

    Response:
        Excel file download.
    """
    try:
        config = current_app.config.get('APP_CONFIG', {})
        mongo_uri, mongo_db, mongo_collection = get_mongo_config(config)

        if not mongo_uri or not mongo_db:
            log_error("MongoDB not configured", endpoint="/data/export/excel")
            return jsonify({"error": "MongoDB not configured"}), 500

        from pymongo import MongoClient
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        client = MongoClient(mongo_uri)
        db = client[mongo_db]
        collection = db[mongo_collection]

        # Check for specific document
        doc_id = request.args.get('doc_id')
        if doc_id:
            doc = find_document_by_id(collection, doc_id)
            client.close()
            if not doc:
                return jsonify({"error": "Document not found"}), 404
            results = [serialize_document(doc)]
        else:
            limit = min(int(request.args.get('limit', 1000)), 10000)
            cursor = collection.find({}).sort("created_at", -1).limit(limit)
            results = [serialize_document(doc) for doc in cursor]
            client.close()

        # Create Excel workbook
        wb = Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Create Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary_headers = ["PDF File", "Total Clauses", "Total Clause Types", "Total Fields", "API Calls", "Created At"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for row, result in enumerate(results, 2):
            ws_summary.cell(row=row, column=1, value=result.get("pdf_file", "")).border = border
            ws_summary.cell(row=row, column=2, value=result.get("total_clauses", 0)).border = border
            ws_summary.cell(row=row, column=3, value=result.get("total_clause_types", 0)).border = border
            ws_summary.cell(row=row, column=4, value=result.get("total_fields", 0)).border = border
            ws_summary.cell(row=row, column=5, value=result.get("openai_api_calls", 0)).border = border
            ws_summary.cell(row=row, column=6, value=result.get("created_at", "")).border = border

        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 18
        ws_summary.column_dimensions['D'].width = 15
        ws_summary.column_dimensions['E'].width = 12
        ws_summary.column_dimensions['F'].width = 22

        # Create Clauses sheet
        ws_clauses = wb.create_sheet("Clauses")
        clause_headers = ["PDF File", "Clause Index", "Clause Type", "Type ID", "Confidence", "Clause Text"]
        for col, header in enumerate(clause_headers, 1):
            cell = ws_clauses.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        clause_row = 2
        for result in results:
            pdf_file = result.get("pdf_file", "")
            for clause_group in result.get("clauses", []):
                clause_type = clause_group.get("type", "")
                clause_type_id = clause_group.get("type_id", "")

                # Handle new grouped format (has 'values' array) or old format
                if "values" in clause_group:
                    for clause in clause_group.get("values", []):
                        ws_clauses.cell(row=clause_row, column=1, value=pdf_file).border = border
                        ws_clauses.cell(row=clause_row, column=2, value=clause.get("clause_index", "")).border = border
                        ws_clauses.cell(row=clause_row, column=3, value=clause_type).border = border
                        ws_clauses.cell(row=clause_row, column=4, value=clause_type_id).border = border
                        ws_clauses.cell(row=clause_row, column=5, value=clause.get("confidence", 0)).border = border

                        clause_text = clause.get("text", "")
                        if len(clause_text) > 32000:
                            clause_text = clause_text[:32000] + "..."
                        cell = ws_clauses.cell(row=clause_row, column=6, value=clause_text)
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True)
                        clause_row += 1
                else:
                    # Old flat format (backward compatibility)
                    ws_clauses.cell(row=clause_row, column=1, value=pdf_file).border = border
                    ws_clauses.cell(row=clause_row, column=2, value=clause_group.get("clause_index", "")).border = border
                    ws_clauses.cell(row=clause_row, column=3, value=clause_group.get("type", "")).border = border
                    ws_clauses.cell(row=clause_row, column=4, value=clause_group.get("type_id", "")).border = border
                    ws_clauses.cell(row=clause_row, column=5, value=clause_group.get("confidence", 0)).border = border

                    clause_text = clause_group.get("text", "")
                    if len(clause_text) > 32000:
                        clause_text = clause_text[:32000] + "..."
                    cell = ws_clauses.cell(row=clause_row, column=6, value=clause_text)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True)
                    clause_row += 1

        ws_clauses.column_dimensions['A'].width = 30
        ws_clauses.column_dimensions['B'].width = 12
        ws_clauses.column_dimensions['C'].width = 25
        ws_clauses.column_dimensions['D'].width = 25
        ws_clauses.column_dimensions['E'].width = 12
        ws_clauses.column_dimensions['F'].width = 80

        # Create Fields sheet
        ws_fields = wb.create_sheet("Fields")
        field_headers = ["PDF File", "Field Name", "Field ID", "Values", "Clause Indices"]
        for col, header in enumerate(field_headers, 1):
            cell = ws_fields.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        field_row = 2
        for result in results:
            pdf_file = result.get("pdf_file", "")
            for field in result.get("fields", []):
                ws_fields.cell(row=field_row, column=1, value=pdf_file).border = border
                ws_fields.cell(row=field_row, column=2, value=field.get("field_name", "")).border = border
                ws_fields.cell(row=field_row, column=3, value=field.get("field_id", "")).border = border

                values = field.get("values", [])
                values_str = ", ".join(str(v) for v in values) if values else ""
                ws_fields.cell(row=field_row, column=4, value=values_str).border = border

                indices = field.get("clause_indices", [])
                indices_str = ", ".join(str(i) for i in indices) if indices else ""
                ws_fields.cell(row=field_row, column=5, value=indices_str).border = border

                field_row += 1

        ws_fields.column_dimensions['A'].width = 30
        ws_fields.column_dimensions['B'].width = 30
        ws_fields.column_dimensions['C'].width = 25
        ws_fields.column_dimensions['D'].width = 50
        ws_fields.column_dimensions['E'].width = 20

        # Save to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"lease_classification_{timestamp}.xlsx"

        log_success("Excel export completed", endpoint="/data/export/excel", count=len(results))

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except ImportError as e:
        log_error("Required library not installed", endpoint="/data/export/excel", error=str(e))
        return jsonify({"error": f"Required library not installed: {str(e)}"}), 500
    except Exception as e:
        log_error("Excel export failed", endpoint="/data/export/excel", error=str(e))
        return jsonify({"error": str(e)}), 500


@data_bp.route('/data/export/pdf', methods=['GET'])
def export_pdf():
    """
    Export classification data as PDF file.

    Query Parameters:
        - limit: Maximum number of records to export (default: 100)
        - doc_id: Export specific document by ID (optional)

    Response:
        PDF file download.
    """
    try:
        config = current_app.config.get('APP_CONFIG', {})
        mongo_uri, mongo_db, mongo_collection = get_mongo_config(config)

        if not mongo_uri or not mongo_db:
            log_error("MongoDB not configured", endpoint="/data/export/pdf")
            return jsonify({"error": "MongoDB not configured"}), 500

        from pymongo import MongoClient
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_LEFT, TA_CENTER

        client = MongoClient(mongo_uri)
        db = client[mongo_db]
        collection = db[mongo_collection]

        # Check for specific document
        doc_id = request.args.get('doc_id')
        if doc_id:
            doc = find_document_by_id(collection, doc_id)
            client.close()
            if not doc:
                return jsonify({"error": "Document not found"}), 404
            results = [serialize_document(doc)]
        else:
            limit = min(int(request.args.get('limit', 100)), 1000)
            cursor = collection.find({}).sort("created_at", -1).limit(limit)
            results = [serialize_document(doc) for doc in cursor]
            client.close()

        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            spaceBefore=15
        )
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            leading=12
        )
        small_style = ParagraphStyle(
            'Small',
            parent=styles['Normal'],
            fontSize=8,
            leading=10
        )
        cell_style = ParagraphStyle(
            'CellText',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
            wordWrap='CJK'
        )
        field_cell_style = ParagraphStyle(
            'FieldCellText',
            parent=styles['Normal'],
            fontSize=8,
            leading=10,
            wordWrap='CJK'
        )

        elements = []

        # Title
        elements.append(Paragraph("Lease Clause Classification Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
        elements.append(Spacer(1, 20))

        # Summary Section
        elements.append(Paragraph("Summary", heading_style))

        summary_data = [["PDF File", "Clauses", "Types", "Fields", "API Calls"]]
        for result in results:
            summary_data.append([
                result.get("pdf_file", "")[:40],
                str(result.get("total_clauses", 0)),
                str(result.get("total_clause_types", 0)),
                str(result.get("total_fields", 0)),
                str(result.get("openai_api_calls", 0))
            ])

        summary_table = Table(summary_data, colWidths=[180, 50, 50, 50, 60])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E8EEF7')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Extracted Fields Section
        elements.append(Paragraph("Extracted Fields", heading_style))

        for result in results:
            pdf_name = result.get("pdf_file", "Unknown")
            fields = result.get("fields", [])

            if fields:
                elements.append(Paragraph(f"<b>{pdf_name}</b>", normal_style))
                elements.append(Spacer(1, 5))

                field_data = [[
                    Paragraph("<b>Field Name</b>", field_cell_style),
                    Paragraph("<b>Values</b>", field_cell_style)
                ]]
                for field in fields:
                    field_name = field.get("field_name", "")
                    values = field.get("values", [])
                    values_str = ", ".join(str(v) for v in values)
                    values_str = values_str.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    field_name = field_name.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    field_data.append([
                        Paragraph(field_name, field_cell_style),
                        Paragraph(values_str, field_cell_style)
                    ])

                field_table = Table(field_data, colWidths=[150, 350])
                field_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5B9BD5')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#DEEBF7')),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                elements.append(field_table)
                elements.append(Spacer(1, 15))

        # Clauses Section (new page)
        elements.append(PageBreak())
        elements.append(Paragraph("Classified Clauses", heading_style))

        for result in results:
            pdf_name = result.get("pdf_file", "Unknown")
            clause_groups = result.get("clauses", [])

            if clause_groups:
                # Flatten clauses for display
                total_clauses = 0
                flat_clauses = []

                for clause_group in clause_groups:
                    if "values" in clause_group:
                        clause_type = clause_group.get("type", "")
                        for clause in clause_group.get("values", []):
                            flat_clauses.append({
                                "clause_index": clause.get("clause_index", ""),
                                "type": clause_type,
                                "confidence": clause.get("confidence", 0),
                                "text": clause.get("text", "")
                            })
                            total_clauses += 1
                    else:
                        flat_clauses.append(clause_group)
                        total_clauses += 1

                elements.append(Paragraph(f"<b>{pdf_name}</b> ({total_clauses} clauses)", normal_style))
                elements.append(Spacer(1, 5))

                clause_data = [[
                    Paragraph("<b>#</b>", cell_style),
                    Paragraph("<b>Type</b>", cell_style),
                    Paragraph("<b>Confidence</b>", cell_style),
                    Paragraph("<b>Text Preview</b>", cell_style)
                ]]
                for clause in flat_clauses[:50]:
                    clause_text = clause.get("text", "")
                    if len(clause_text) > 300:
                        clause_text = clause_text[:300] + "..."
                    clause_text = clause_text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    clause_type = clause.get("type", "")[:30]

                    clause_data.append([
                        Paragraph(str(clause.get("clause_index", "")), cell_style),
                        Paragraph(clause_type, cell_style),
                        Paragraph(f"{clause.get('confidence', 0):.2%}", cell_style),
                        Paragraph(clause_text, cell_style)
                    ])

                clause_table = Table(clause_data, colWidths=[25, 90, 50, 340])
                clause_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                    ('ALIGN', (2, 1), (2, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E2EFDA')),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                elements.append(clause_table)

                if total_clauses > 50:
                    elements.append(Paragraph(f"<i>... and {total_clauses - 50} more clauses</i>", small_style))

                elements.append(Spacer(1, 15))

        doc.build(elements)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"lease_classification_{timestamp}.pdf"

        log_success("PDF export completed", endpoint="/data/export/pdf", count=len(results))

        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except ImportError as e:
        log_error("Required library not installed", endpoint="/data/export/pdf", error=str(e))
        return jsonify({"error": f"Required library not installed: {str(e)}"}), 500
    except Exception as e:
        log_error("PDF export failed", endpoint="/data/export/pdf", error=str(e))
        return jsonify({"error": str(e)}), 500
